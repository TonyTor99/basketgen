from __future__ import annotations

import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable

import numpy as np
import openpyxl
import pandas as pd


ODDS_LABELS = {
    "П1", "X", "П2", "ТБ", "ТМ", "ФОРА", "Инд.тотал 1", "Инд.тотал 2",
    "КФ ТБ", "КФ ТМ", "КФ ФОРА1", "КФ ФОРА2", "КФ ИТБ1", "КФ ИТМ1",
    "КФ ИТБ2", "КФ ИТМ2", "1-я полов. ТБ", "1-я полов. КФ ТБ",
    "1-я полов. ТМ", "1-я полов. КФ ТМ", "ФОРА1", "ФОРА2", "ИТБ1",
    "ИТМ1", "ИТБ2", "ИТМ2",
}

EXCLUDE_FEATURE_TERMS = [
    "3 четв",
    "4 четв",
    "итоговый счет",
    "текущий банк",
    "max ставка",
    "размер ставки",
]

PREMATCH_LAST_SCAN_TM_COLUMN = "Прематчевые КФы (последнее сканирование)::ТМ"
PREMATCH_FIRST_SCAN_TM_COLUMN = "Прематчевые КФы (первое сканирование)::ТМ"
PREMATCH_TM_FEATURE_NAME = "ТМ (прематч: последнее vs первое)"

DEFAULT_FEATURE_PRIORITY = [
    "Голы",
    "Фолы",
    "2-х очковые",
    "3-х очковые",
    "Штраф броски",
    "Реализац атак, %",
    "1 четв",
    "2 четв",
    "Ост таймаут",
    PREMATCH_TM_FEATURE_NAME,
]


@dataclass(slots=True)
class FeaturePair:
    name: str
    home_col: str
    away_col: str
    completeness: float


@dataclass(slots=True)
class HeaderInfo:
    names: list[str]
    index_by_name: dict[str, int]


@dataclass(slots=True)
class FilterDiagnostics:
    rows_initial: int
    rows_after_bot_filter: int
    rows_after_signal_time: int
    rows_after_min_odds: int
    rows_after_max_odds: int
    rows_after_required_fields: int
    rows_final: int

    def as_dict(self) -> dict[str, int]:
        return {
            "rows_initial": int(self.rows_initial),
            "rows_after_bot_filter": int(self.rows_after_bot_filter),
            "rows_after_signal_time": int(self.rows_after_signal_time),
            "rows_after_min_odds": int(self.rows_after_min_odds),
            "rows_after_max_odds": int(self.rows_after_max_odds),
            "rows_after_required_fields": int(self.rows_after_required_fields),
            "rows_final": int(self.rows_final),
        }


@dataclass(slots=True)
class DatasetPreparationResult:
    df: pd.DataFrame
    codes: np.ndarray
    diagnostics: FilterDiagnostics


class ExcelFormatError(RuntimeError):
    pass


class DatasetPreparationError(RuntimeError):
    def __init__(self, message: str, diagnostics: FilterDiagnostics):
        super().__init__(message)
        self.diagnostics = diagnostics


BASE_DATASET_COLUMNS = [
    "Дата",
    "Чемпионат",
    'Команда "Хозяева"',
    'Команда "Гости"',
    "Прогноз",
    "Коэффициент",
    "Результат",
    "Прибыль/убыток",
    "Бот",
    "Время сигнала",
]

REQUIRED_BASE_FIELDS = [
    "Дата",
    "Чемпионат",
    'Команда "Хозяева"',
    'Команда "Гости"',
    "Прогноз",
    "Коэффициент",
    "Результат",
    "Прибыль/убыток",
]

SUMMARY_KEY_FIELDS = [
    "Дата",
    "Бот",
    "Время сигнала",
    "Коэффициент",
    "Результат",
    "Прибыль/убыток",
    'Команда "Хозяева"',
    'Команда "Гости"',
]


def build_header_info(path: Path) -> HeaderInfo:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    row0 = [v for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    row1 = [v for v in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
    wb.close()

    groups: list[str | None] = []
    current_group: str | None = None
    for value in row0:
        if value not in (None, ""):
            current_group = str(value).strip()
        groups.append(current_group)

    base_names: list[str] = []
    for idx, (group, sub) in enumerate(zip(groups, row1)):
        sub_name = None if sub is None else str(sub).strip()
        group_name = None if group is None else str(group).strip()

        if sub_name:
            base = sub_name
            if group_name and group_name != sub_name and sub_name in ODDS_LABELS:
                base = f"{group_name}::{sub_name}"
        else:
            base = group_name or f"col_{idx}"
        base_names.append(base)

    counts: dict[str, int] = {}
    for name in base_names:
        counts[name] = counts.get(name, 0) + 1

    seen: dict[str, int] = {}
    final_names: list[str] = []
    for name in base_names:
        seen[name] = seen.get(name, 0) + 1
        final_names.append(name if counts[name] == 1 else f"{name}__{seen[name]}")

    return HeaderInfo(names=final_names, index_by_name={name: i for i, name in enumerate(final_names)})



def _pair_name(header: str) -> tuple[str, str] | None:
    match = re.match(r"^(.*?)(?:\sД)(.*)$", header)
    if not match:
        return None
    base, suffix = match.groups()
    return base, suffix



def detect_feature_pairs(path: Path, header_info: HeaderInfo | None = None, sample_rows: int = 8000) -> list[FeaturePair]:
    header_info = header_info or build_header_info(path)
    pairs_meta: list[tuple[str, str, str]] = []
    for header in header_info.names:
        pair = _pair_name(header)
        if not pair:
            continue
        base, suffix = pair
        away_col = f"{base} Г{suffix}"
        if away_col not in header_info.index_by_name:
            continue
        feature_name = f"{base}{suffix}".strip()
        lowered = feature_name.lower()
        if any(term in lowered for term in EXCLUDE_FEATURE_TERMS):
            continue
        pairs_meta.append((feature_name, header, away_col))

    if (
        PREMATCH_LAST_SCAN_TM_COLUMN in header_info.index_by_name
        and PREMATCH_FIRST_SCAN_TM_COLUMN in header_info.index_by_name
    ):
        pairs_meta.append(
            (
                PREMATCH_TM_FEATURE_NAME,
                PREMATCH_LAST_SCAN_TM_COLUMN,
                PREMATCH_FIRST_SCAN_TM_COLUMN,
            )
        )

    pairs_meta = list(dict.fromkeys(pairs_meta))
    if not pairs_meta:
        return []

    sample_cols = [home for _, home, _ in pairs_meta] + [away for _, _, away in pairs_meta]
    sample_df = read_subset(path, header_info, sample_cols, nrows=sample_rows)

    features: list[FeaturePair] = []
    for name, home, away in pairs_meta:
        home_vals = pd.to_numeric(sample_df[home], errors="coerce")
        away_vals = pd.to_numeric(sample_df[away], errors="coerce")
        completeness = float((~(home_vals.isna() | away_vals.isna())).mean())
        label = "Очки на момент сигнала" if name == "Голы" else name
        features.append(FeaturePair(name=label, home_col=home, away_col=away, completeness=completeness))

    def feature_key(item: FeaturePair) -> tuple[int, float, str]:
        try:
            priority = DEFAULT_FEATURE_PRIORITY.index(item.name)
        except ValueError:
            priority = 999
        return (priority, -item.completeness, item.name)

    return sorted(features, key=feature_key)



def read_subset(path: Path, header_info: HeaderInfo, columns: Iterable[str], nrows: int | None = None) -> pd.DataFrame:
    selected = list(dict.fromkeys(columns))
    missing = [col for col in selected if col not in header_info.index_by_name]
    if missing:
        raise ExcelFormatError(f"Не найдены колонки: {', '.join(missing)}")

    indexed = [(header_info.index_by_name[col], col) for col in selected]
    indexed_sorted = sorted(indexed, key=lambda item: item[0])
    usecols = [idx for idx, _ in indexed_sorted]
    names_for_read = [name for _, name in indexed_sorted]

    df = pd.read_excel(
        path,
        header=1,
        names=names_for_read,
        usecols=usecols,
        nrows=nrows,
        engine="openpyxl",
    )
    return df[selected]


def _filled_mask(series: pd.Series) -> pd.Series:
    mask = series.notna()
    if pd.api.types.is_object_dtype(series) or pd.api.types.is_string_dtype(series):
        stripped = series.astype(str).str.strip()
        mask &= stripped.ne("")
        mask &= stripped.str.lower().ne("nan")
    return mask


def build_quick_summary(path: Path, header_info: HeaderInfo | None = None) -> dict[str, object]:
    header_info = header_info or build_header_info(path)
    df = read_subset(path, header_info, BASE_DATASET_COLUMNS)

    rows_total = int(len(df))
    bots = df["Бот"]
    bot_counts = (
        bots[_filled_mask(bots)]
        .astype(str)
        .str.strip()
        .value_counts(dropna=True)
    )

    signal_series = pd.to_numeric(df["Время сигнала"], errors="coerce").dropna()
    signal_values: list[int | float] = []
    for value in sorted(signal_series.unique().tolist()):
        as_float = float(value)
        signal_values.append(int(as_float) if as_float.is_integer() else round(as_float, 2))

    odds_series = pd.to_numeric(df["Коэффициент"], errors="coerce")
    odds_min = round(float(odds_series.min()), 3) if odds_series.notna().any() else None
    odds_max = round(float(odds_series.max()), 3) if odds_series.notna().any() else None

    key_fields: list[dict[str, object]] = []
    for field in SUMMARY_KEY_FIELDS:
        mask = _filled_mask(df[field])
        filled = int(mask.sum())
        percent = round((filled / rows_total * 100.0) if rows_total else 0.0, 2)
        key_fields.append({"field": field, "filled": filled, "percent": percent})

    return {
        "rows_total": rows_total,
        "bot_counts": [{"value": str(name), "count": int(count)} for name, count in bot_counts.items()],
        "signal_time_values": signal_values,
        "odds_min": odds_min,
        "odds_max": odds_max,
        "key_fields": key_fields,
    }


def format_filter_diagnostics(diagnostics: FilterDiagnostics) -> list[str]:
    values = diagnostics.as_dict()
    return [
        f"Всего матчей: {values['rows_initial']}",
        f"После фильтра по боту: {values['rows_after_bot_filter']}",
        f"После фильтра по времени сигнала: {values['rows_after_signal_time']}",
        f"После мин. коэффициента: {values['rows_after_min_odds']}",
        f"После макс. коэффициента: {values['rows_after_max_odds']}",
        f"После удаления строк с пустыми базовыми полями: {values['rows_after_required_fields']}",
        f"Итог: {values['rows_final']}",
    ]


def explain_empty_dataset(diagnostics: FilterDiagnostics) -> str:
    if diagnostics.rows_after_bot_filter == 0:
        return "После фильтра по боту не осталось матчей."
    if diagnostics.rows_after_signal_time == 0:
        return "После фильтра по времени сигнала не осталось матчей."
    if diagnostics.rows_after_min_odds == 0 or diagnostics.rows_after_max_odds == 0:
        return "После фильтра по коэффициентам не осталось матчей."
    if diagnostics.rows_after_required_fields == 0:
        return "После удаления строк с пустыми базовыми полями не осталось матчей."
    if diagnostics.rows_final == 0:
        return "После расчета ставок не осталось матчей."
    return "После фильтров не осталось матчей."



def estimate_stake(profit: float, odds: float, outcome: str) -> float:
    if pd.isna(profit) or pd.isna(odds):
        return math.nan
    if outcome == "loss":
        return abs(float(profit))
    if outcome == "win":
        return abs(float(profit)) / max(float(odds) - 1.0, 1e-9)
    if outcome == "push":
        return max(abs(float(profit)), 1.0)
    return math.nan



def normalize_outcome(result_value: object, profit: float) -> str:
    value = str(result_value).strip().lower()
    if "выиг" in value:
        return "win"
    if "проиг" in value:
        return "loss"
    if "возв" in value or "расч" in value:
        return "push"
    if pd.notna(profit):
        if profit > 0:
            return "win"
        if profit < 0:
            return "loss"
    return "push"



def build_codes(df: pd.DataFrame, pairs: list[FeaturePair]) -> np.ndarray:
    codes = np.zeros((len(df), len(pairs)), dtype=np.int8)
    for idx, pair in enumerate(pairs):
        home_vals = pd.to_numeric(df[pair.home_col], errors="coerce").to_numpy(dtype=float)
        away_vals = pd.to_numeric(df[pair.away_col], errors="coerce").to_numpy(dtype=float)
        codes[:, idx] = np.where(
            np.isnan(home_vals) | np.isnan(away_vals),
            0,
            np.where(home_vals > away_vals, 1, np.where(home_vals < away_vals, 2, 0)),
        )
    return codes



def prepare_dataset_detailed(
    path: Path,
    header_info: HeaderInfo,
    selected_pairs: list[FeaturePair],
    bot_filter: str = "",
    signal_time: int | None = 20,
    min_odds: float | None = None,
    max_odds: float | None = None,
    stage_callback: Callable[[str], None] | None = None,
) -> DatasetPreparationResult:
    feature_cols = [pair.home_col for pair in selected_pairs] + [pair.away_col for pair in selected_pairs]

    if stage_callback:
        stage_callback("reading_excel")
    df = read_subset(path, header_info, BASE_DATASET_COLUMNS + feature_cols)

    if stage_callback:
        stage_callback("preparing_dataset")
    df["Дата"] = pd.to_datetime(df["Дата"], dayfirst=True, errors="coerce")
    df["Коэффициент"] = pd.to_numeric(df["Коэффициент"], errors="coerce")
    df["Прибыль/убыток"] = pd.to_numeric(df["Прибыль/убыток"], errors="coerce")
    df["Время сигнала"] = pd.to_numeric(df["Время сигнала"], errors="coerce")

    for pair in selected_pairs:
        df[pair.home_col] = pd.to_numeric(df[pair.home_col], errors="coerce")
        df[pair.away_col] = pd.to_numeric(df[pair.away_col], errors="coerce")

    if stage_callback:
        stage_callback("applying_filters")
    rows_initial = int(len(df))

    if bot_filter:
        mask = df["Бот"].astype(str).str.contains(bot_filter, case=False, regex=False, na=False)
        df = df[mask]
    rows_after_bot_filter = int(len(df))

    if signal_time is not None:
        df = df[df["Время сигнала"].eq(signal_time)]
    rows_after_signal_time = int(len(df))

    if min_odds is not None:
        df = df[df["Коэффициент"] >= min_odds]
    rows_after_min_odds = int(len(df))

    if max_odds is not None:
        df = df[df["Коэффициент"] <= max_odds]
    rows_after_max_odds = int(len(df))

    df = df[_filled_mask(df["Дата"])]
    for field in REQUIRED_BASE_FIELDS:
        df = df[_filled_mask(df[field])]
    rows_after_required_fields = int(len(df))

    df["outcome"] = [normalize_outcome(res, pnl) for res, pnl in zip(df["Результат"], df["Прибыль/убыток"])]
    df["stake"] = [estimate_stake(pnl, odds, outcome) for pnl, odds, outcome in zip(df["Прибыль/убыток"], df["Коэффициент"], df["outcome"])]
    df = df.replace([np.inf, -np.inf], np.nan)
    df = df.dropna(subset=["stake"])
    df = df.reset_index(drop=True)
    rows_final = int(len(df))

    diagnostics = FilterDiagnostics(
        rows_initial=rows_initial,
        rows_after_bot_filter=rows_after_bot_filter,
        rows_after_signal_time=rows_after_signal_time,
        rows_after_min_odds=rows_after_min_odds,
        rows_after_max_odds=rows_after_max_odds,
        rows_after_required_fields=rows_after_required_fields,
        rows_final=rows_final,
    )
    if rows_final == 0:
        raise DatasetPreparationError(explain_empty_dataset(diagnostics), diagnostics)

    if stage_callback:
        stage_callback("encoding_features")
    codes = build_codes(df, selected_pairs)
    return DatasetPreparationResult(df=df, codes=codes, diagnostics=diagnostics)


def prepare_dataset(
    path: Path,
    header_info: HeaderInfo,
    selected_pairs: list[FeaturePair],
    bot_filter: str = "",
    signal_time: int | None = 20,
    min_odds: float | None = None,
    max_odds: float | None = None,
) -> tuple[pd.DataFrame, np.ndarray]:
    result = prepare_dataset_detailed(
        path=path,
        header_info=header_info,
        selected_pairs=selected_pairs,
        bot_filter=bot_filter,
        signal_time=signal_time,
        min_odds=min_odds,
        max_odds=max_odds,
    )
    return result.df, result.codes
